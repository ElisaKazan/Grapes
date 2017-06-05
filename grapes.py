from openpyxl import load_workbook
import sys

# To run script:
# python grapes.py <name-of-file>

def main():
    # Get name of excel document
    if len(sys.argv) < 2:
        # Error message & exit
        print("Usage: python grapes.py <excel-file-name>")
        sys.exit()
    excel_name = sys.argv[1]
    print("Document name: {}".format(excel_name))

    # Open excel document
    excel_workbook = load_workbook(excel_name)

    # Create grades dictionary
    grades_dict = {}

    # Loop over sheets
    for sheet in excel_workbook:
        print("Sheet - {}".format(sheet.title))

        final_grade = 0

        # Loop over rows (starting at 3)
        for row in sheet.iter_rows(min_row=3, min_col=1, max_col=5):
            if row[0].value == None:
                break
            percentage = get_percentage(row[3].value)
            part_of_grade = float(row[2].value) * percentage
            final_grade += part_of_grade
            print("\tAssignment {} - {:.2f}%".format(row[0].value, part_of_grade))

            # Save part of grade
            row[4].value = part_of_grade

        # Get Grade Received (ex: 90)
        print("FINAL GRADE in {} so far -> {:.2f}%\n".format(sheet.title, final_grade))

    # Save Worksheet
    excel_workbook.save(excel_name)

def get_percentage(grade):
    if grade == None:
        return 0
    elif isinstance(grade, (int,float)):
        if grade <= 1:
            return float(grade)
        elif grade <= 100:
            return float(grade)/100
        else:
            print("Error: Not a valid grade")
            sys.exit()
    elif '/' in grade:
        nums = grade.split('/')
        return float(nums[0])/float(nums[1])
    else:
        print("Error: Something went wrong")


if __name__ == '__main__':
    # executes only if run as script
    main()